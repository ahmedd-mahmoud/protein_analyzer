import os
import re
import time
import json
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from Bio import SeqIO

def get_ncbi_data(protein_id):
    """Fetches the locus tag and protein description from NCBI."""
    print(f"  -> Querying NCBI for protein ID: {protein_id}")
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params = {
        "db": "protein",
        "id": protein_id,
        "rettype": "gb",
        "retmode": "text"
    }
    try:
        response = requests.get(base_url, params=params)
        response.raise_for_status()
        content = response.text

        # Use regular expressions to find the locus_tag
        locus_tag_match = re.search(r'/locus_tag="([^"]+)"', content)
        locus_tag = locus_tag_match.group(1) if locus_tag_match else "Not Found"

        # Find the definition (protein name)
        definition_match = re.search(r"DEFINITION\s+(.+)", content)
        definition = definition_match.group(1).strip() if definition_match else "Not Found"
        
        return locus_tag, definition
    except requests.exceptions.RequestException as e:
        print(f"    [ERROR] Could not fetch data from NCBI for {protein_id}. Error: {e}")
        return "Error", "Error"


def get_alphafold_data(locus_tag: str):
    """
    Fetches key data from the AlphaFold API using a locus tag.

    This function uses a two-step process:
    1. Uses the AlphaFold search API to find the UniProt ID corresponding to the locus tag.
    2. Uses that UniProt ID to fetch the full prediction data.
    
    Returns a tuple: (prediction_data_dict, uniprot_id) or (None, None) on failure.
    """
    if not locus_tag:
        return None, None

    # --- Step 1: Search for the locus tag to find the UniProt ID ---
    print(f"\n  -> Searching AlphaFold for locus tag: '{locus_tag}'...")
    url = "https://alphafold.ebi.ac.uk/api/search"
    params = {
        "q": f"(text:*{locus_tag} OR text:{locus_tag}*)",
        "type": "main",
        "start": 0,
        "rows": 20
    }
    
    uniprot_id = None
    try:
        response = requests.get(url, params=params)
        response.raise_for_status() # Raises an HTTPError for bad responses (4xx or 5xx)
        
        search_results = response.json()

        if not search_results.get("docs"):
            print(f"     [INFO] No search results found for locus tag '{locus_tag}'.")
            return None, None
            
        # We assume the first hit is the most relevant one for a specific locus tag.
        first_hit = search_results["docs"][0]
        uniprot_id = first_hit.get("uniprotAccession")
        
        if not uniprot_id:
            print(f"     [ERROR] Search successful, but could not extract UniProt ID for '{locus_tag}'.")
            return None, None
            
        print(f"     Found corresponding UniProt ID: {uniprot_id}")

    except requests.exceptions.RequestException as e:
        print(f"     [ERROR] Could not perform search on AlphaFold for '{locus_tag}'. Error: {e}")
        return None, None
    except json.JSONDecodeError:
        print(f"     [ERROR] Failed to decode JSON response from AlphaFold search for '{locus_tag}'.")
        return None, None

    # --- Step 2: Fetch the prediction data using the found UniProt ID ---
    print(f"  -> Querying AlphaFold for prediction data for {uniprot_id}...")
    prediction_url = f"https://alphafold.ebi.ac.uk/api/prediction/{uniprot_id}"
    try:
        response = requests.get(prediction_url)
        if response.status_code == 200:
            print(f"     AlphaFold data retrieved successfully for {uniprot_id}.")
            # The API returns a list, we want the first (and only) item
            return response.json()[0], uniprot_id
        else:
            print(f"     [INFO] No AlphaFold entry found for {uniprot_id} (Status code: {response.status_code})")
            return None, None
    except requests.exceptions.RequestException as e:
        print(f"     [ERROR] Could not fetch prediction data from AlphaFold for {uniprot_id}. Error: {e}")
        return None, None

def get_species_type(uniprot_id):
    """Fetches protein Clusters members number from AlphaFold."""
    print(f"  -> Querying for protein Clusters members number: {uniprot_id}")
    base_url = f"https://alphafold.ebi.ac.uk/api/cluster/members/{uniprot_id}"
    params = {
        "cluster_flag": "AFDB50/MMseqs2",
        "records": 5,
        "start": 0,
        "sort_direction": "DESC",
        "sort_column": "averagePlddt"
    }
    try:
        response = requests.get(base_url, params=params)
        response.raise_for_status()
        result = response.json()
        
        return result['clusterTotal']
    except requests.exceptions.RequestException as e:
        print(f"    [ERROR] Could not fetch  protein Clusters members number for {uniprot_id}. Error: {e}")
        return "Error"

def download_file(url, save_path):
    """Downloads a file from a URL to a specified path."""
    try:
        print(f"  -> Downloading file from: {url}")
        response = requests.get(url, stream=True)
        response.raise_for_status()
        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        print(f"     -> Successfully saved to: {os.path.basename(save_path)}")
        return True
    except requests.exceptions.RequestException as e:
        print(f"     [ERROR] Failed to download file. Error: {e}")
        return False

# --- MAIN SCRIPT ---

def main():
    """Main function to run the analysis pipeline."""
    # --- 1. SETUP YOUR FILE PATHS HERE ---
    # IMPORTANT: Use forward slashes '/' in your paths, even on Windows.
    fasta_input_file = "D:/Career Learning/Randoms/file2.fasta"
    main_output_directory = "D:/Career Learning/Randoms/output"

    print("--- Starting Protein Analysis Pipeline ---")

    os.makedirs(main_output_directory, exist_ok=True)
    all_proteins_data = []
    protein_counter = 0

    try:
        fasta_sequences = list(SeqIO.parse(fasta_input_file, "fasta"))
        print(f"Found {len(fasta_sequences)} proteins in the FASTA file.")
    except FileNotFoundError:
        print(f"[FATAL ERROR] The input FASTA file was not found at: {fasta_input_file}")
        print("Please check the path and try again.")
        return

    for record in fasta_sequences:
        protein_counter += 1
        protein_id = record.id
        print(f"\n--- Processing Protein {protein_counter}/{len(fasta_sequences)}: {protein_id} ---")

        # Create a dedicated folder for this protein's files
        protein_folder_path = os.path.join(main_output_directory, str(protein_counter))
        os.makedirs(protein_folder_path, exist_ok=True)

        locus_tag, ncbi_protein_name = get_ncbi_data(protein_id)
        
        # Initialize AlphaFold data with default values
        aa_length = 'N/A'
        plddt = 'N/A'
        species_value = '' # Default to empty
        alpha_missense_value = '' # Default to empty

        uniprot_id = "Not Found"
        af_data = None
        if locus_tag not in ["Not Found", "Error"]:
            af_data, uniprot_id = get_alphafold_data(locus_tag)
        
        if af_data:
            aa_length = af_data.get('sequenceEnd', 'N/A')
            plddt = af_data.get('globalMetricValue', 'N/A')
            pdb_url = af_data.get('pdbUrl')

            # Process Species data
            try:
                # Navigate through the nested dictionary to get the hit count
                similar_proteins_count = get_species_type(uniprot_id)
                print(f"     Found {similar_proteins_count} similar proteins (AFDB50/MMseqs2).")
                if similar_proteins_count >= 100:
                    species_value = "General"
                elif similar_proteins_count < 10:
                    species_value = "Specific"
                # Otherwise, it remains empty as initialized
            except (KeyError, TypeError):
                print("     [INFO] Could not find similar proteins count in AlphaFold data.")

            # Process Alpha missense data
            try:
                # Check if the 'alphaMissense' list exists and is not empty
                missense_annotations = af_data.get('amAnnotationsUrl')
                if missense_annotations and isinstance(missense_annotations, list) and len(missense_annotations) > 0:
                    alpha_missense_value = "Yes"
                    print("     Found Alpha Missense annotations.")
                else:
                    print("     No Alpha Missense annotations found.")
            except (KeyError, TypeError):
                print("     [INFO] Could not find Alpha Missense data.")

            # Download the PDB file if a URL was found
            if pdb_url:
                pdb_filename = f"{uniprot_id}.pdb"
                pdb_save_path = os.path.join(protein_folder_path, pdb_filename)
                download_file(pdb_url, pdb_save_path)
        
        # Data for DALI, MTM, Interpro etc. are placeholders for you to fill in.
        protein_data = {
            'Folder': protein_counter,
            'Protein ID': protein_id,
            'Gene ID': locus_tag,
            'Uniport ID': uniprot_id,
            'AA': aa_length,
            'PLDDT': plddt, # TODO: get plddt from protein domain or average of its domains plddts  
            'D.ID': '',
            'D.Z SCORE': '',
            'D.RMSD': '',
            'D.NAME': '',
            'M.ID': '',
            'M.IM': '',
            'M.RSMD': '',
            'M.NAME': '',
            'Interpro': 'MANUAL ENTRY: Run InterProScan and summarize results here.',
            'NCBI': ncbi_protein_name,
            'PDB E.value': 'MANUAL ENTRY: From AlphaFold website.',
            'Species': species_value, 
            'Alpha missense': alpha_missense_value
        }
        all_proteins_data.append(protein_data)
        
        # A short delay to be respectful to the servers
        time.sleep(1)

    print("\n--- All proteins processed. Generating Excel report... ---")
    df = pd.DataFrame(all_proteins_data)

    # Reorder columns to match your example
    column_order = [
        'Folder', 'Protein ID', 'Gene ID', 'Uniport ID', 'AA', 'PLDDT',
        'D.ID', 'D.Z SCORE', 'D.RMSD', 'D.NAME', 'M.ID', 'M.IM', 'M.RSMD',
        'M.NAME', 'Interpro', 'NCBI', 'PDB E.value', 'Species', 'Alpha missense'
    ]
    df = df[column_order]

    excel_output_path = os.path.join(main_output_directory, "analysis_results.xlsx")
    df.to_excel(excel_output_path, index=False)

    wb = load_workbook(excel_output_path)
    ws = wb.active

    # Define highlight style (red fill if plddt < 70)
    highlight_fill = PatternFill(start_color="f1160a", end_color="f1160a", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2):
        plddt_cell = row[5]
        if isinstance(plddt_cell.value, (int, float)) and plddt_cell.value < 70:
            plddt_cell.fill = highlight_fill

    wb.save(excel_output_path)

    print(f"--- Pipeline Finished! ---")
    print(f"Results saved to: {excel_output_path}")
    print(f"PDB files and other data are organized in numbered folders inside: {main_output_directory}")

if __name__ == "__main__":
    main()