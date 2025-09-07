"""Service for file operations including FASTA reading and Excel generation."""

import logging
import os
from pathlib import Path
from typing import List, Optional

import pandas as pd
import requests
from Bio import SeqIO
from Bio.SeqRecord import SeqRecord
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ..core.models import ProteinData
from ..shared.constants import (
    DEFAULT_CHUNK_SIZE,
    DEFAULT_OUTPUT_FILENAME,
    EXCEL_COLUMN_ORDER,
    PLDDT_HIGHLIGHT_THRESHOLD,
    PDB_EXTENSION,
)
from ..shared.exceptions import FileProcessingError
from ..shared.utils import create_protein_folder, sanitize_filename

logger = logging.getLogger(__name__)


class FileService:
    """Service for handling file operations."""

    def __init__(self, chunk_size: int = DEFAULT_CHUNK_SIZE):
        """
        Initialize the file service.

        Args:
            chunk_size (int): Chunk size for file downloads.
        """
        self.chunk_size = chunk_size

    def read_fasta_file(self, file_path: str) -> List[SeqRecord]:
        """
        Read and parse a FASTA file.

        Args:
            file_path (str): Path to the FASTA file.

        Returns:
            List[SeqRecord]: List of sequence records from the file.

        Raises:
            FileProcessingError: If the file cannot be read or parsed.
        """
        try:
            logger.info(f"Reading FASTA file: {file_path}")
            
            # Try different encodings if UTF-8 fails
            encodings = ['utf-8', 'latin-1', 'cp1252']
            sequences = None
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as handle:
                        sequences = list(SeqIO.parse(handle, "fasta"))
                    break
                except UnicodeDecodeError:
                    continue
                    
            if sequences is None:
                raise FileProcessingError(f"Could not decode file with any supported encoding")
                
            if not sequences:
                raise FileProcessingError(f"No valid sequences found in FASTA file")
                
            logger.info(f"Successfully read {len(sequences)} sequences from FASTA file")
            return sequences
            
        except FileNotFoundError as e:
            error_msg = f"FASTA file not found: {file_path}"
            logger.error(error_msg)
            raise FileProcessingError(error_msg) from e
        except Exception as e:
            error_msg = f"Failed to parse FASTA file {file_path}: {str(e)}"
            logger.error(error_msg)
            raise FileProcessingError(error_msg) from e

    def download_pdb_file(
        self, url: str, output_directory: str, uniprot_id: str
    ) -> bool:
        """
        Download a PDB file from a given URL.

        Args:
            url (str): URL to download the PDB file from.
            output_directory (str): Directory to save the file in.
            uniprot_id (str): UniProt ID for filename generation.

        Returns:
            bool: True if download was successful, False otherwise.
        """
        if not url or not uniprot_id:
            logger.warning("Invalid URL or UniProt ID provided for PDB download")
            return False

        filename = sanitize_filename(f"{uniprot_id}{PDB_EXTENSION}")
        save_path = os.path.join(output_directory, filename)

        # Skip if file already exists and is not empty
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            logger.info(f"PDB file already exists: {filename}")
            return True

        try:
            logger.info(f"Downloading PDB file from: {url}")
            
            response = requests.get(url, stream=True, timeout=30)
            response.raise_for_status()

            # Check content type
            content_type = response.headers.get('content-type', '')
            if 'text/plain' not in content_type and 'application/octet-stream' not in content_type:
                logger.warning(f"Unexpected content type for PDB file: {content_type}")

            with open(save_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=self.chunk_size):
                    if chunk:
                        f.write(chunk)

            # Verify file was downloaded and is not empty
            if os.path.getsize(save_path) == 0:
                os.remove(save_path)
                logger.error(f"Downloaded PDB file is empty: {filename}")
                return False

            logger.info(f"Successfully saved PDB file to: {filename}")
            return True

        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to download PDB file from {url}: {e}")
            return False
        except OSError as e:
            logger.error(f"Failed to save PDB file to {save_path}: {e}")
            return False
        except Exception as e:
            logger.error(f"Unexpected error downloading PDB file: {e}")
            return False

    def create_excel_report(
        self, protein_data_list: List[ProteinData], output_directory: str
    ) -> str:
        """
        Create an Excel report from protein data.

        Args:
            protein_data_list (List[ProteinData]): List of protein data objects.
            output_directory (str): Directory to save the Excel file.

        Returns:
            str: Path to the created Excel file.

        Raises:
            FileProcessingError: If the Excel file cannot be created.
        """
        try:
            logger.info(f"Creating Excel report with {len(protein_data_list)} proteins")

            # Convert protein data to dictionaries
            data_dicts = [protein.to_dict() for protein in protein_data_list]

            # Create DataFrame with proper column ordering
            df = pd.DataFrame(data_dicts)
            df = df[EXCEL_COLUMN_ORDER]

            # Save to Excel
            excel_path = os.path.join(output_directory, DEFAULT_OUTPUT_FILENAME)
            df.to_excel(excel_path, index=False)

            # Apply formatting
            self._apply_excel_formatting(excel_path)

            logger.info(f"Excel report created successfully: {excel_path}")
            return excel_path

        except Exception as e:
            error_msg = f"Failed to create Excel report: {str(e)}"
            logger.error(error_msg)
            raise FileProcessingError(error_msg) from e

    def _apply_excel_formatting(self, excel_path: str) -> None:
        """
        Apply formatting to the Excel file (highlight low PLDDT values).

        Args:
            excel_path (str): Path to the Excel file to format.
        """
        try:
            logger.info("Applying Excel formatting")
            
            workbook = load_workbook(excel_path)
            worksheet = workbook.active

            # Define highlight style for low PLDDT values
            highlight_fill = PatternFill(
                start_color="f1160a", end_color="f1160a", fill_type="solid"
            )

            # Find PLDDT column dynamically
            plddt_column_index = None
            for idx, cell in enumerate(worksheet[1]):  # Header row
                if cell.value == "PLDDT":
                    plddt_column_index = idx
                    break

            if plddt_column_index is None:
                logger.warning("PLDDT column not found in Excel file")
                return

            for row in worksheet.iter_rows(min_row=2):  # Skip header row
                if plddt_column_index < len(row):
                    plddt_cell = row[plddt_column_index]
                    
                    # Check if the value is numeric and below threshold
                    try:
                        if plddt_cell.value and plddt_cell.value != "N/A":
                            plddt_value = float(plddt_cell.value)
                            if plddt_value < PLDDT_HIGHLIGHT_THRESHOLD:
                                plddt_cell.fill = highlight_fill
                    except (ValueError, TypeError):
                        # Skip non-numeric values
                        continue

            workbook.save(excel_path)
            logger.info("Excel formatting applied successfully")

        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting: {e}")
            # Don't raise an exception here as the core functionality still works

    def create_protein_output_folder(
        self, base_output_dir: str, protein_counter: int
    ) -> str:
        """
        Create a dedicated folder for a protein's output files.

        Args:
            base_output_dir (str): Base output directory.
            protein_counter (int): Protein counter for folder naming.

        Returns:
            str: Path to the created folder.

        Raises:
            FileProcessingError: If the folder cannot be created.
        """
        try:
            return create_protein_folder(base_output_dir, protein_counter)
        except OSError as e:
            error_msg = f"Failed to create protein folder {protein_counter}: {str(e)}"
            logger.error(error_msg)
            raise FileProcessingError(error_msg) from e

    def validate_output_directory(self, directory_path: str) -> bool:
        """
        Validate and create output directory if needed.

        Args:
            directory_path (str): Path to the output directory.

        Returns:
            bool: True if directory is valid/created, False otherwise.
        """
        try:
            Path(directory_path).mkdir(parents=True, exist_ok=True)
            return True
        except (OSError, PermissionError) as e:
            logger.error(f"Failed to create/validate output directory {directory_path}: {e}")
            return False

    def get_file_size(self, file_path: str) -> Optional[int]:
        """
        Get the size of a file in bytes.

        Args:
            file_path (str): Path to the file.

        Returns:
            Optional[int]: File size in bytes, None if file doesn't exist.
        """
        try:
            return os.path.getsize(file_path)
        except OSError:
            return None

    def cleanup_temp_files(self, directory: str, pattern: str = "*.tmp") -> None:
        """
        Clean up temporary files in a directory.

        Args:
            directory (str): Directory to clean up.
            pattern (str): File pattern to match for cleanup.
        """
        try:
            path = Path(directory)
            for temp_file in path.glob(pattern):
                try:
                    temp_file.unlink()
                    logger.debug(f"Cleaned up temporary file: {temp_file}")
                except OSError as e:
                    logger.warning(f"Failed to remove temporary file {temp_file}: {e}")
        except Exception as e:
            logger.warning(f"Error during cleanup of {directory}: {e}")